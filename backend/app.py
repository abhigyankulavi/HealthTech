# backend/app.py

from flask import Flask, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Define the Excel file path
EXCEL_FILE = "backend/patient_records.xlsx"  # Store the Excel file in the backend directory

# Create the Excel file with headers if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    # Add headers
    sheet.append(["Name", "Age", "Gender", "Contact", "Address", "Allergies", "Chronic Conditions"])
    workbook.save(EXCEL_FILE)

# Define a route to handle form submissions
@app.route('/submit_record', methods=['POST'])
def submit_record():
    # Get form data
    name = request.form['name']
    age = request.form['age']
    gender = request.form['gender']
    contact = request.form['contact']
    address = request.form.get('address', '')
    allergies = request.form.get('allergies', '')
    chronic_conditions = request.form.get('chronic_conditions', '')

    # Open the Excel file and append new row
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([name, age, gender, contact, address, allergies, chronic_conditions])
    workbook.save(EXCEL_FILE)

    return "Record added successfully"
